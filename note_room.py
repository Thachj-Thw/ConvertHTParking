from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, Color, Alignment, Border, Side, borders

START_USER = 5
START_CARD = 6
MAX_ROW = 10000
NOTE = "F"
ROOM = "C"
CARD_NO = "C"
CARD_ID = "B"
CARD_USER = "W"
ORDINAL = "A"
CARD_TYPE = "D"
CARD_LOCK = "G"

def read_user(sheet):
    result = {}
    for row in range(START_USER, MAX_ROW):
        str_row = str(row)
        card_no = sheet[CARD_USER + str_row].value
        room = sheet[ROOM + str_row].value
        if not card_no and not room:
            break
        result[card_no] = room
    return result

def set_cell(cell: Cell, value: str|int, bold: bool = False, font_color: str = "00000000", size: int = 12, alignment: str = "center"):
        font = Font(bold=bold, color=font_color, size=size)
        alig = Alignment(horizontal=alignment)
        border_type = Side(border_style=borders.BORDER_THIN)
        bor = Border(top=border_type, right=border_type, bottom=border_type, left=border_type)
        cell.font = font
        cell.alignment = alig
        #cell.border = bor
        cell.value = value

def fill_card(sheet, user_data):
    result = {}
    card = {}
    for row in range(START_CARD, MAX_ROW):
        str_row = str(row)
        card_no = sheet[CARD_NO + str_row].value
        card_id = sheet[CARD_ID + str_row].value
        if not card_no and card_id:
            break
        card[card_no] = card_id
    for card_no in user_data.keys():
        if card_no in card:
            result[card_no] = {
                "id": card[card_no],
                "room": user_data[card_no]
            }
    
    return result

def save_card(file_name, data, lock=False):
    wb = load_workbook("ConvertHTParking/templates/BaseCard.xlsx")
    sheet = wb.active
    for i, card_no in enumerate(data.keys()):
        str_row = str(START_CARD + i)
        set_cell(sheet[ORDINAL + str_row], i+1)
        set_cell(sheet[CARD_ID + str_row], data[card_no]["id"])
        set_cell(sheet[CARD_NO + str_row], card_no)
        set_cell(sheet[CARD_TYPE + str_row], 2)
        set_cell(sheet[NOTE + str_row], data[card_no]["room"])
        if lock:
            set_cell(sheet[CARD_LOCK + str_row], "x")
    wb.save(file_name)

user = "Output/Dang ky khach hang.xlsx"
card = "Output/Dang ky the.xlsx"
lock = "Output/Dang ky the_lock.xlsx"

wb_user = load_workbook(user)
wb_card = load_workbook(card)
wb_lock = load_workbook(lock)

user_sheet = wb_user.active
card_sheet = wb_card.active
lock_sheet = wb_lock.active

user_data = read_user(user_sheet)
card_use = fill_card(card_sheet, user_data)
card_lock = fill_card(lock_sheet, user_data)
save_card("Output/Dang ky the note.xlsx", card_use)
save_card("Output/Dang ky the note lock.xlsx", card_lock, lock=True)
