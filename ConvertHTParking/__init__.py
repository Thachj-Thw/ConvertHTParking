from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import borders
from openpyxl.cell import Cell
import module #module-thw
from .structs import InputForm, OutputUserForm, OutputCardForm, DataStruct
from .autoformat import format_card_code, format_vehicle_plate_motor
from .convert import convert8to10
import re
import os


path = module.Path(__file__)



class HTParking:

    card_form = OutputCardForm(
        ordinal     = "A",
        card_id     = "B",
        card_no     = "C",
        card_type   = "D",
        vehicle     = "E",
        status      = "G",
        start       = 6
    )

    user_form = OutputUserForm(
        ordinal         = "A",
        name            = "B",
        address         = "C",
        user_id         = "E",
        end_time        = "K",
        vehicle         = "P",
        card_no         = "W",
        vehicle_plate   = "R",
        vehicle_type    = "Q",
        start           = 5
    )


class KzParking:

    user_form = InputForm(
        card_no         = "B",
        card_id         = "C",
        vehicle         = "D",
        vehicle_plate   = "E",
        end_time        = "F",
        status          = "G",
        user_id         = "C",
        name            = "I",
        address         = "J",
        start           = 9
    )

    def __init__(self, convert_vehicle: dict, vehicle_month: list, prioritize: list,
                 fix_duplicate_card_method: callable,
                 end_row: int = 10000,
                 emty_count: int = 1,
                 format_card_method: callable = format_card_code, 
                 format_vehicle_plate_method: callable = lambda p, _: format_vehicle_plate_motor(p), 
                 convert_card_id: callable = convert8to10):
        self.data = {}
        self.emty_count = emty_count
        self.fix_card = fix_duplicate_card_method
        self.vehicle_month = vehicle_month
        self.prioritize = prioritize
        self.convert_vehicle = convert_vehicle
        self.end_row = end_row
        self.format_card = format_card_method
        self.format_vehicle_plate = format_vehicle_plate_method
        self.convert_card_id = convert_card_id

    def _read_excel(self, file_name: str, form: InputForm, pass_vehicle: str = None) -> list:
        wb = load_workbook(file_name)
        sheet = wb.active
        result = []
        emty_counter = 0
        for row in range(form.start, self.end_row):
            str_row = str(row)
            vehicle = sheet[form.vehicle + str_row].value
            card_no = sheet[form.card_no + str_row].value
            card_id = sheet[form.card_id + str_row].value
            vehicle_plate = sheet[form.vehicle_plate + str_row].value
            end_time = sheet[form.end_time + str_row].value
            status = sheet[form.status + str_row].value
            name = sheet[form.name + str_row].value
            address = sheet[form.address + str_row].value
            if not vehicle and not card_no and not card_id and not vehicle_plate and not end_time and not status and not name and not address:
                emty_counter += 1
                if emty_counter > self.emty_count:
                    break
            if vehicle in self.convert_vehicle and self.convert_vehicle[vehicle] == pass_vehicle:
                continue
            result.append({
                "vehicle": vehicle,
                "card_no": card_no,
                "card_id": card_id,
                "vehicle_plate": vehicle_plate,
                "end_time": end_time,
                "status": status,
                "name": name,
                "address": address
            })
        return result
    
    def _split_data(self, lst_data: list) -> tuple[dict, dict]:
        card_data = {}
        user_data = {}
        def add_data(data: dict, card_no: str, vehicle: str):
            card_data[card_id] = {
                "card_no": card_no,
                "vehicle": vehicle,
                "status": data["status"]
            }
            if card_data[card_id]["vehicle"] in self.vehicle_month:
                user_data[card_id] = {
                    "card": card_data[card_id],
                    "name": data["name"],
                    "vehicle_plate": self.format_vehicle_plate(data["vehicle_plate"], data),
                    "end_time": data["end_time"],
                    "address": data["address"].rstrip().lstrip()
                }
            else:
                if card_id in user_data:
                    user_data.pop(card_id)

        for d in lst_data:
            if not d["card_id"] or not d["card_id"].isnumeric():
                continue
            if not d["card_no"]:
                continue
            if not d["vehicle"]:
                continue
            vehicle = self.convert_vehicle[d["vehicle"]]
            if vehicle in self.vehicle_month and not d["name"]:
                continue
            card_id = self.convert_card_id(d["card_id"])
            card_no = self.format_card(d["card_no"])
            
            if card_id not in card_data:
                add_data(d, card_no, vehicle)
            else:
                current_status = card_data[card_id]["status"]
                new_status = d["status"]
                if current_status != "Hoạt động" and new_status == "Hoạt động":
                    add_data(d, card_no, vehicle)
                    continue
                if current_status == "Hoạt động" and new_status != "Hoạt động":
                    continue
                current_is_month = card_data[card_id]["vehicle"] in self.vehicle_month
                new_is_month = vehicle in self.vehicle_month
                if current_is_month and not new_is_month:
                    continue
                if not current_is_month and new_is_month:
                    add_data(d, card_no, vehicle)
                    continue
                current = card_data[card_id]["vehicle"]
                new = d["vehicle"]
                if not self._is_current_priority_than(current, new, self.prioritize):
                    add_data(d, card_no, vehicle)
                
        return card_data, user_data


    def _is_current_priority_than(self, current: str, new: str, base: list) -> bool:
        try:
            current_index = base.index(current)
        except:
            return False
        try:
            new_index = base.index(new)
        except:
            return True
        if current_index < new_index:
            return True
        return False
    
    def _fix_card_duplicate(self, card_data: dict, user_data: dict) -> None:
        valid_card_no = []
        for card_id in card_data.keys():
            card_no = card_data[card_id]["card_no"]
            if card_no in valid_card_no:
                print("Trung ma the", card_no)
                card_no = self.fix_card(card_data[card_id]["card_no"], card_data[card_id], valid_card_no)
                card_data[card_id]["card_no"] = card_no
                if card_id in user_data:
                    user_data[card_id]["card"] = card_data[card_id]
            valid_card_no.append(card_no)

    def _set_cell(self, cell: Cell, value: str|int, bold: bool = False, font_color: str = "00000000", size: int = 12, alignment: str = "center"):
        font = Font(bold=bold, color=font_color, size=size)
        alig = Alignment(horizontal=alignment)
        border_type = Side(border_style=borders.BORDER_THIN)
        bor = Border(top=border_type, right=border_type, bottom=border_type, left=border_type)
        cell.font = font
        cell.alignment = alig
        #cell.border = bor
        cell.value = value
    
    def _save_card_excel(self, card_data: dict, save_as: str, form: OutputCardForm) -> None:
        card_file = load_workbook(filename=path.source.join("templates", "BaseCard.xlsx"))
        card_sheet = card_file.active
        ordinal = 0
        for card_id in card_data.keys():
            str_row = str(form.start + ordinal)
            ordinal += 1
            self._set_cell(card_sheet[form.ordinal + str_row], ordinal)
            self._set_cell(card_sheet[form.card_id + str_row], card_id)
            self._set_cell(card_sheet[form.card_no + str_row], card_data[card_id]["card_no"])
            if card_data[card_id]["vehicle"] in self.vehicle_month:
                self._set_cell(card_sheet[form.card_type + str_row], 2)
            else:
                self._set_cell(card_sheet[form.card_type + str_row], 1)
                self._set_cell(card_sheet[form.vehicle + str_row], card_data[card_id]["vehicle"])
        card_file.save(save_as)


    def _save_user_excel(self, user_data: dict, save_as: str, form: OutputUserForm) -> None:
        user_file = load_workbook(filename=path.source.join("templates", "BaseUser.xlsx"))
        user_sheet = user_file.active
        ordinal = 0
        for user_id in user_data:
            str_row = str(form.start + ordinal)
            ordinal += 1
            self._set_cell(user_sheet[form.ordinal + str_row], ordinal)
            self._set_cell(user_sheet[form.name + str_row], user_data[user_id]["name"])
            self._set_cell(user_sheet[form.user_id + str_row], user_id)
            self._set_cell(user_sheet[form.address + str_row], user_data[user_id]["address"])
            self._set_cell(user_sheet[form.vehicle + str_row], user_data[user_id]["card"]["vehicle"])
            self._set_cell(user_sheet[form.end_time + str_row], user_data[user_id]["end_time"])
            self._set_cell(user_sheet[form.card_no + str_row], user_data[user_id]["card"]["card_no"])
            vehicle_plate = user_data[user_id]["vehicle_plate"]
            if len(vehicle_plate) > 15:
                self._set_cell(user_sheet[form.vehicle_type + str_row], vehicle_plate)
            else:
                self._set_cell(user_sheet[form.vehicle_plate + str_row], vehicle_plate)
        user_file.save(save_as)
        
    def _split_card_lock(self, card_data: dict) -> tuple[dict, dict]:
        card_use = {}
        card_lock = {}
        for card_id in card_data.keys():
            if card_data[card_id]["status"] == "Hoạt động":
                card_use[card_id] = card_data[card_id]
            else:
                card_lock[card_id] = card_data[card_id]
        return card_lock, card_use

    def convert(self, lst_file: list, save_card: str, save_user: str, pass_index: int = -1, pass_vehicle: str = None) -> None:
        all_data = []
        for idx, file in enumerate(lst_file):
            if idx == pass_index:
                all_data += self._read_excel(file, KzParking.user_form, pass_vehicle)
            else:
                all_data += self._read_excel(file, KzParking.user_form, None)

        card_data, user_data = self._split_data(all_data)
        self._fix_card_duplicate(card_data, user_data)
        card_lock, card_use = self._split_card_lock(card_data)
        self._save_card_excel(card_use, save_card, HTParking.card_form)
        self._save_user_excel(user_data, save_user, HTParking.user_form)
        tail = os.path.splitext(save_card)[1]
        self._save_card_excel(card_lock, save_card.replace(tail, "_lock" + tail), HTParking.card_form)
